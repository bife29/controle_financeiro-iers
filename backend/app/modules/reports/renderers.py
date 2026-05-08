"""Geração de relatórios em PDF (reportlab) e XLSX (openpyxl).

Padrão das funções: recebem (title, subtitle, columns, rows, totals?) e retornam bytes.
- columns: lista de tuplas (header, key, align?) onde align ∈ {"left","right","center"}
- rows: lista de dicts
- totals: dict opcional {key: valor} renderizado como linha final em destaque
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)


def _fmt_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return str(v)


def render_xlsx(
    title: str,
    subtitle: str,
    columns: list[tuple[str, str]] | list[tuple[str, str, str]],
    rows: Iterable[dict],
    totals: dict | None = None,
    sections: list[dict] | None = None,
) -> bytes:
    """Gera XLSX. Se `sections` for fornecido, ignora `rows/totals` e gera
    múltiplos blocos no mesmo sheet (cada section: {title, columns, rows, totals?}).
    """
    wb = Workbook()
    ws = wb.active
    # openpyxl proíbe : \ / ? * [ ] no título do sheet (limite 31 chars)
    safe_title = "".join("-" if c in r':\/?*[]' else c for c in (title or ""))
    ws.title = safe_title[:31] or "Relatório"

    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    total_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 1
    ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=14)
    row_idx += 1
    if subtitle:
        ws.cell(row=row_idx, column=1, value=subtitle).font = Font(italic=True, color="475569")
        row_idx += 1
    ws.cell(row=row_idx, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=9, color="94A3B8")
    row_idx += 2

    blocks = sections if sections else [{
        "title": None,
        "columns": columns,
        "rows": list(rows),
        "totals": totals,
    }]

    for block in blocks:
        cols = block["columns"]
        if block.get("title"):
            ws.cell(row=row_idx, column=1, value=block["title"]).font = Font(bold=True, size=12, color="1E3A8A")
            row_idx += 1
        # header
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=col[0])
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        # rows
        for r in block["rows"]:
            for ci, col in enumerate(cols, start=1):
                key = col[1]
                align = col[2] if len(col) > 2 else "left"
                val = r.get(key)
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="center")
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
            row_idx += 1
        # totals
        bt = block.get("totals")
        if bt:
            for ci, col in enumerate(cols, start=1):
                key = col[1]
                align = col[2] if len(col) > 2 else "left"
                val = bt.get(key, "" if ci > 1 else "TOTAL")
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="center")
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
            row_idx += 1
        row_idx += 1  # spacer entre blocos

    # auto-width simples
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, min(40, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_pdf(
    title: str,
    subtitle: str,
    columns: list[tuple[str, str]] | list[tuple[str, str, str]],
    rows: Iterable[dict],
    totals: dict | None = None,
    sections: list[dict] | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14, spaceAfter=2, textColor=colors.HexColor("#1E3A8A"))
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#475569"), spaceAfter=2)
    meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#94A3B8"), spaceAfter=8)
    section_h = ParagraphStyle("section_h", parent=styles["Heading2"], fontSize=11, textColor=colors.HexColor("#1E3A8A"), spaceBefore=6, spaceAfter=4)

    story: list = []
    story.append(Paragraph(title, h1))
    if subtitle:
        story.append(Paragraph(subtitle, sub))
    story.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", meta))

    blocks = sections if sections else [{
        "title": None,
        "columns": columns,
        "rows": list(rows),
        "totals": totals,
    }]

    align_map = {"left": "LEFT", "right": "RIGHT", "center": "CENTER"}

    for block in blocks:
        cols = block["columns"]
        if block.get("title"):
            story.append(Paragraph(block["title"], section_h))

        header = [c[0] for c in cols]
        data = [header]
        for r in block["rows"]:
            data.append([_fmt_value(r.get(c[1])) for c in cols])
        if block.get("totals"):
            t = block["totals"]
            data.append([
                _fmt_value(t.get(c[1], "TOTAL" if i == 0 else ""))
                for i, c in enumerate(cols)
            ])

        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if block.get("totals") else -1),
             [colors.white, colors.HexColor("#F8FAFC")]),
        ])
        for ci, col in enumerate(cols):
            align = col[2] if len(col) > 2 else "left"
            ts.add("ALIGN", (ci, 1), (ci, -1), align_map.get(align, "LEFT"))
        if block.get("totals"):
            ts.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF3C7"))
            ts.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")

        table = Table(data, repeatRows=1)
        table.setStyle(ts)
        story.append(table)
        story.append(Spacer(1, 8))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def streaming_response(content: bytes, filename: str, fmt: str):
    """Helper para retornar como StreamingResponse."""
    from fastapi.responses import Response
    if fmt == "xlsx":
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        media = "application/pdf"
    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
